"""Structured BOM ingestion parser for CSV, JSON, and XLSX formats.

Sprint C: Accepts real-world BOM files with messy, inconsistent data and
converts them into ParsedBOMItem models that can be fed to the TEaaS solver.

Key capabilities:
- Fuzzy column matching with configurable alias maps
- Material/property extraction from description fields using the canonical
  vocabulary from fact_vocabulary.py
- Material percentage computation by weight and value
- Resilient to real-world messiness: empty rows, section headers, bad data
"""

from __future__ import annotations

import csv
import json
import io
import logging
import re
import unicodedata
from typing import Any, Dict, List, Optional, Set, Tuple

from pydantic import BaseModel, ConfigDict, Field

from potatobacon.tariff.fact_vocabulary import (
    ENTAILMENTS,
    MATERIAL_CHAPTER_TOKENS,
    get_entailed_tokens,
)
from potatobacon.tariff.product_schema import (
    MaterialBreakdown,
    ProductCategory,
    ProductSpecModel,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column alias map: canonical_name -> set of known aliases
# ---------------------------------------------------------------------------
COLUMN_ALIASES: Dict[str, List[str]] = {
    "part_id": [
        "part_id", "partid", "part_number", "partnumber", "part_no",
        "partno", "part no", "part no.", "sku", "sku_id", "skuid",
        "item_id", "itemid", "item_number", "itemnumber", "item_no",
        "itemno", "item no", "item no.", "id", "pn", "p/n",
    ],
    "description": [
        "description", "desc", "name", "part_description",
        "partdescription", "part_name", "partname", "item_description",
        "itemdescription", "item_name", "itemname", "product",
        "product_name", "productname", "product_description",
        "productdescription", "component", "component_name",
    ],
    "material": [
        "material", "materials", "composition", "material_type",
        "materialtype", "mat", "mat_type", "mattype", "substance",
        "raw_material", "rawmaterial",
    ],
    "weight_kg": [
        "weight_kg", "weightkg", "weight", "mass", "weight_g",
        "weightg", "mass_kg", "masskg", "net_weight", "netweight",
        "gross_weight", "grossweight", "wt", "wt_kg",
    ],
    "value_usd": [
        "value_usd", "valueusd", "unit_cost", "unitcost",
        "unit cost", "unit_cost_usd", "price", "unit_price",
        "unitprice", "cost", "value", "amount", "ext_cost",
        "extcost", "unit cost (usd)", "unit_cost_(usd)",
    ],
    "origin_country": [
        "origin_country", "origincountry", "country", "coo",
        "country_of_origin", "countryoforigin", "origin", "source_country",
        "sourcecountry", "made_in", "madein", "mfg_country",
        "mfgcountry", "country of origin",
    ],
    "hts_code": [
        "hts_code", "htscode", "hs_code", "hscode", "tariff_code",
        "tariffcode", "hts", "hs", "hts_number", "htsnumber",
        "hs_number", "hsnumber", "classification", "tariff",
        "hts code", "hs code", "tariff code",
    ],
    "quantity": [
        "quantity", "qty", "count", "units", "pcs", "pieces",
        "qty_per_unit", "qtyperunit", "quantity_per_unit",
    ],
}

# Build reverse lookup: normalized alias -> canonical name
_ALIAS_LOOKUP: Dict[str, str] = {}
for _canonical, _aliases in COLUMN_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_LOOKUP[_alias] = _canonical

# Required columns (at least one must map)
REQUIRED_COLUMNS = {"description"}
RECOMMENDED_COLUMNS = {"part_id", "material", "value_usd", "origin_country"}


def _normalize_header(header: str) -> str:
    """Normalize a column header for fuzzy matching.

    Lowercases, strips whitespace, removes underscores/hyphens/dots,
    collapses multiple spaces, and strips non-ASCII.
    """
    # Normalize unicode
    text = unicodedata.normalize("NFKD", header)
    # Lowercase
    text = text.lower().strip()
    # Replace underscores, hyphens, dots with spaces
    text = re.sub(r"[_\-.]", " ", text)
    # Remove non-alphanumeric except spaces and parens
    text = re.sub(r"[^a-z0-9\s()]", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text).strip()
    return text


def match_columns(headers: List[str]) -> Tuple[Dict[str, str], List[str], List[str]]:
    """Match raw column headers to canonical column names.

    Returns:
        (mapping, matched, unmatched) where mapping is {raw_header: canonical_name},
        matched is list of canonical names found, unmatched is list of raw headers
        that couldn't be matched.
    """
    mapping: Dict[str, str] = {}
    matched_canonical: Set[str] = set()
    unmatched: List[str] = []

    for raw in headers:
        normalized = _normalize_header(raw)
        # Also try without spaces (for things like "partid")
        no_spaces = normalized.replace(" ", "")

        canonical = _ALIAS_LOOKUP.get(normalized) or _ALIAS_LOOKUP.get(no_spaces)
        if canonical and canonical not in matched_canonical:
            mapping[raw] = canonical
            matched_canonical.add(canonical)
        else:
            unmatched.append(raw)

    return mapping, sorted(matched_canonical), unmatched


# ---------------------------------------------------------------------------
# Material extraction from descriptions using canonical vocabulary
# ---------------------------------------------------------------------------
# Material keywords -> fact token
_DESCRIPTION_MATERIAL_KEYWORDS: Dict[str, str] = {
    "steel": "material_steel",
    "stainless steel": "material_steel",
    "stainless": "material_steel",
    "iron": "material_steel",
    "aluminum": "material_aluminum",
    "aluminium": "material_aluminum",
    "copper": "material_copper",
    "brass": "material_copper",
    "bronze": "material_copper",
    "plastic": "material_plastic",
    "abs": "material_plastic",
    "pvc": "material_plastic",
    "polyethylene": "material_plastic",
    "polypropylene": "material_plastic",
    "nylon": "material_synthetic",
    "polyester": "material_synthetic",
    "rubber": "material_rubber",
    "silicone": "material_rubber",
    "leather": "material_leather",
    "wood": "material_wood",
    "wooden": "material_wood",
    "glass": "material_glass",
    "ceramic": "material_ceramic",
    "textile": "material_textile",
    "fabric": "material_textile",
    "cotton": "material_textile",
    "silk": "material_textile",
    "wool": "material_textile",
    "linen": "material_textile",
}

# Product type keywords -> fact token
_DESCRIPTION_PRODUCT_KEYWORDS: Dict[str, str] = {
    "bolt": "is_fastener",
    "screw": "is_fastener",
    "nut": "is_fastener",
    "washer": "is_fastener",
    "rivet": "is_fastener",
    "fastener": "is_fastener",
    "cable": "product_type_cable",
    "wire": "product_type_cable",
    "harness": "product_type_cable",
    "connector": "product_type_cable",
    "pcb": "product_type_electronics",
    "circuit board": "product_type_electronics",
    "battery": "product_type_battery",
    "motor": "product_type_machinery",
    "pump": "product_type_machinery",
    "engine": "product_type_machinery",
    "compressor": "product_type_machinery",
    "shoe": "product_type_footwear",
    "boot": "product_type_footwear",
    "sandal": "product_type_footwear",
    "footwear": "product_type_footwear",
    "chair": "product_type_furniture",
    "desk": "product_type_furniture",
    "table": "product_type_furniture",
    "furniture": "product_type_furniture",
    "garment": "product_type_apparel",
    "shirt": "product_type_apparel",
    "jacket": "product_type_apparel",
    "trouser": "product_type_apparel",
}

# Surface treatment keywords
_SURFACE_TREATMENT_KEYWORDS: Dict[str, str] = {
    "galvanized": "surface_treatment_galvanized",
    "zinc-plated": "surface_treatment_galvanized",
    "zinc plated": "surface_treatment_galvanized",
    "chrome": "surface_treatment_chrome",
    "chrome-plated": "surface_treatment_chrome",
    "nickel-plated": "surface_treatment_nickel",
    "painted": "surface_treatment_painted",
    "anodized": "surface_treatment_anodized",
    "coated": "has_coating",
    "laminated": "has_lamination",
    "plated": "has_plating",
    "insulated": "is_insulated",
}


def extract_facts_from_description(description: str) -> Dict[str, bool]:
    """Extract canonical fact tokens from a BOM row description.

    Uses regex + keyword matching against the canonical vocabulary from
    fact_vocabulary.py so extracted facts are compatible with guard tokens.

    Args:
        description: Free-text product/part description.

    Returns:
        Dict of {fact_token: True} for all detected facts.
    """
    lower = description.lower()
    facts: Dict[str, bool] = {}

    # Sort by length descending to match longer phrases first
    for keyword in sorted(_DESCRIPTION_MATERIAL_KEYWORDS.keys(), key=len, reverse=True):
        if keyword in lower:
            facts[_DESCRIPTION_MATERIAL_KEYWORDS[keyword]] = True

    for keyword in sorted(_DESCRIPTION_PRODUCT_KEYWORDS.keys(), key=len, reverse=True):
        if keyword in lower:
            facts[_DESCRIPTION_PRODUCT_KEYWORDS[keyword]] = True

    for keyword in sorted(_SURFACE_TREATMENT_KEYWORDS.keys(), key=len, reverse=True):
        if keyword in lower:
            facts[_SURFACE_TREATMENT_KEYWORDS[keyword]] = True

    # Add entailed tokens for all discovered facts
    entailed: Dict[str, bool] = {}
    for token in list(facts.keys()):
        for implied in get_entailed_tokens(token):
            entailed[implied] = True
    facts.update(entailed)

    return facts


def _infer_category_from_facts(facts: Dict[str, bool]) -> ProductCategory:
    """Infer ProductCategory from extracted fact tokens."""
    if facts.get("is_fastener") or facts.get("product_type_fastener"):
        return ProductCategory.FASTENER
    if facts.get("product_type_electronics") or facts.get("product_type_battery"):
        return ProductCategory.ELECTRONICS
    if facts.get("product_type_cable"):
        return ProductCategory.ELECTRONICS
    if facts.get("product_type_footwear"):
        return ProductCategory.FOOTWEAR
    if facts.get("product_type_apparel"):
        return ProductCategory.APPAREL_TEXTILE
    if facts.get("material_textile"):
        return ProductCategory.TEXTILE
    if facts.get("product_type_furniture"):
        return ProductCategory.FURNITURE
    return ProductCategory.OTHER


# ---------------------------------------------------------------------------
# ParsedBOMItem and parsing logic
# ---------------------------------------------------------------------------
class ParsedBOMItem(BaseModel):
    """A single parsed BOM row, ready for conversion to ProductSpecModel."""

    row_number: int = Field(description="1-based row number in the source file")
    part_id: Optional[str] = None
    description: str
    material: Optional[str] = None
    weight_kg: Optional[float] = None
    value_usd: Optional[float] = None
    origin_country: Optional[str] = None
    hts_code: Optional[str] = None
    quantity: Optional[float] = None
    extracted_facts: Dict[str, bool] = Field(default_factory=dict)
    inferred_category: str = "other"
    warnings: List[str] = Field(default_factory=list)

    model_config = ConfigDict(extra="forbid")

    def to_product_spec(self) -> ProductSpecModel:
        """Convert this parsed BOM item to a ProductSpecModel for the solver."""
        category = ProductCategory(self.inferred_category)

        materials: List[MaterialBreakdown] = []
        if self.material:
            materials.append(
                MaterialBreakdown(
                    component=self.description[:50] if self.description else "component",
                    material=self.material.lower(),
                )
            )
        else:
            # Infer material from extracted facts
            for fact_key in sorted(self.extracted_facts.keys()):
                if fact_key.startswith("material_") and self.extracted_facts[fact_key]:
                    mat_name = fact_key.replace("material_", "")
                    materials.append(
                        MaterialBreakdown(
                            component=self.description[:50] if self.description else "component",
                            material=mat_name,
                        )
                    )

        return ProductSpecModel(
            product_category=category,
            materials=materials,
            origin_country=self.origin_country,
            import_country="US",
            declared_value_per_unit=self.value_usd,
        )


class SkippedRow(BaseModel):
    """A row that was skipped during parsing, with reason."""

    row_number: int
    reason: str
    raw_data: Optional[Dict[str, Any]] = None

    model_config = ConfigDict(extra="forbid")


class BOMParseResult(BaseModel):
    """Result of parsing a BOM file."""

    items: List[ParsedBOMItem]
    skipped: List[SkippedRow]
    column_mapping: Dict[str, str]
    detected_columns: List[str]
    unmatched_columns: List[str]
    total_rows: int
    warnings: List[str] = Field(default_factory=list)

    model_config = ConfigDict(extra="forbid")


# ---------------------------------------------------------------------------
# Row-level parsing helpers
# ---------------------------------------------------------------------------
def _coerce_float(value: Any) -> Optional[float]:
    """Coerce a value to float, handling currency symbols and commas."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Remove currency symbols and commas
    text = re.sub(r"[$€£¥,]", "", text)
    text = text.strip()
    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def _is_section_header(row: Dict[str, Any]) -> bool:
    """Detect if a row is likely a section header, not an actual BOM line."""
    values = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]
    if len(values) <= 1:
        return True  # Only one field populated → likely a header

    # Check if description looks like a section header (all caps, short, no numbers)
    desc = row.get("description", "")
    if desc and isinstance(desc, str):
        desc = desc.strip()
        if desc.isupper() and len(desc) < 50 and not any(c.isdigit() for c in desc):
            # Looks like "MECHANICAL PARTS" or "ELECTRONICS ASSEMBLY"
            non_empty = sum(
                1 for k, v in row.items()
                if k != "description" and v is not None and str(v).strip()
            )
            if non_empty == 0:
                return True
    return False


def _is_empty_row(row: Dict[str, Any]) -> bool:
    """Check if a row is entirely empty."""
    return all(
        v is None or str(v).strip() == ""
        for v in row.values()
    )


def _parse_row(
    row_data: Dict[str, Any],
    column_mapping: Dict[str, str],
    row_number: int,
) -> Tuple[Optional[ParsedBOMItem], Optional[SkippedRow]]:
    """Parse a single row into a ParsedBOMItem or SkippedRow."""
    # Map raw columns to canonical names
    canonical: Dict[str, Any] = {}
    for raw_key, value in row_data.items():
        mapped = column_mapping.get(raw_key)
        if mapped:
            canonical[mapped] = value

    # Check for empty row
    if _is_empty_row(canonical):
        return None, SkippedRow(row_number=row_number, reason="Empty row")

    # Check for section header
    if _is_section_header(canonical):
        desc = str(canonical.get("description", "")).strip()
        return None, SkippedRow(
            row_number=row_number,
            reason=f"Section header: {desc[:50]}" if desc else "Section header (empty)",
            raw_data={k: str(v)[:100] for k, v in row_data.items() if v},
        )

    # Extract description — required
    description = canonical.get("description")
    if not description or not str(description).strip():
        return None, SkippedRow(
            row_number=row_number,
            reason="Missing description",
            raw_data={k: str(v)[:100] for k, v in row_data.items() if v},
        )
    description = str(description).strip()

    # Extract other fields
    part_id = canonical.get("part_id")
    if part_id is not None:
        part_id = str(part_id).strip() or None

    material = canonical.get("material")
    if material is not None:
        material = str(material).strip() or None

    weight_kg = _coerce_float(canonical.get("weight_kg"))
    value_usd = _coerce_float(canonical.get("value_usd"))
    quantity = _coerce_float(canonical.get("quantity"))

    origin_country = canonical.get("origin_country")
    if origin_country is not None:
        origin_country = str(origin_country).strip().upper() or None

    hts_code = canonical.get("hts_code")
    if hts_code is not None:
        hts_code = str(hts_code).strip() or None

    warnings: List[str] = []

    # Extract facts from description
    extracted_facts = extract_facts_from_description(description)
    # Also extract from material field if present
    if material:
        mat_facts = extract_facts_from_description(material)
        extracted_facts.update(mat_facts)

    # Infer category
    category = _infer_category_from_facts(extracted_facts)

    return ParsedBOMItem(
        row_number=row_number,
        part_id=part_id,
        description=description,
        material=material,
        weight_kg=weight_kg,
        value_usd=value_usd,
        origin_country=origin_country,
        hts_code=hts_code,
        quantity=quantity,
        extracted_facts=extracted_facts,
        inferred_category=category.value,
        warnings=warnings,
    ), None


# ---------------------------------------------------------------------------
# Format-specific parsers
# ---------------------------------------------------------------------------
def parse_csv(content: bytes | str) -> BOMParseResult:
    """Parse a CSV BOM file.

    Args:
        content: Raw CSV bytes or string.

    Returns:
        BOMParseResult with parsed items and metadata.
    """
    if isinstance(content, bytes):
        # Try UTF-8 first, fall back to latin-1
        try:
            text = content.decode("utf-8-sig")  # handles BOM
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    else:
        text = content

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=["No headers found in CSV file"],
        )

    headers = list(reader.fieldnames)
    column_mapping, detected_columns, unmatched_columns = match_columns(headers)

    # Check for required columns
    parse_warnings: List[str] = []
    missing_required = REQUIRED_COLUMNS - set(detected_columns)
    if missing_required:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping=column_mapping,
            detected_columns=detected_columns,
            unmatched_columns=unmatched_columns,
            total_rows=0,
            warnings=[
                f"Missing required column(s): {sorted(missing_required)}. "
                f"Found columns: {headers}. "
                f"Matched: {detected_columns}."
            ],
        )

    missing_recommended = RECOMMENDED_COLUMNS - set(detected_columns)
    if missing_recommended:
        parse_warnings.append(
            f"Recommended columns not found: {sorted(missing_recommended)}"
        )

    items: List[ParsedBOMItem] = []
    skipped: List[SkippedRow] = []
    total_rows = 0

    for row_idx, row in enumerate(reader, start=2):  # row 1 is header
        total_rows += 1
        item, skip = _parse_row(row, column_mapping, row_number=row_idx)
        if item:
            items.append(item)
        elif skip:
            skipped.append(skip)

    return BOMParseResult(
        items=items,
        skipped=skipped,
        column_mapping=column_mapping,
        detected_columns=detected_columns,
        unmatched_columns=unmatched_columns,
        total_rows=total_rows,
        warnings=parse_warnings,
    )


def parse_json(content: bytes | str) -> BOMParseResult:
    """Parse a JSON BOM file.

    Expects either:
    - A JSON array of objects (each object is a BOM row)
    - A JSON object with an "items" key containing an array

    Args:
        content: Raw JSON bytes or string.

    Returns:
        BOMParseResult with parsed items and metadata.
    """
    if isinstance(content, bytes):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    else:
        text = content

    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=[f"Invalid JSON: {exc}"],
        )

    # Handle both array and object-with-items formats
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict) and "items" in data:
        rows = data["items"]
    else:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=["JSON must be an array or an object with an 'items' key"],
        )

    if not rows:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
        )

    # Collect all unique keys from all rows
    all_keys: List[str] = []
    seen: Set[str] = set()
    for row in rows:
        if isinstance(row, dict):
            for k in row.keys():
                if k not in seen:
                    all_keys.append(k)
                    seen.add(k)

    column_mapping, detected_columns, unmatched_columns = match_columns(all_keys)

    parse_warnings: List[str] = []
    missing_required = REQUIRED_COLUMNS - set(detected_columns)
    if missing_required:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping=column_mapping,
            detected_columns=detected_columns,
            unmatched_columns=unmatched_columns,
            total_rows=len(rows),
            warnings=[
                f"Missing required column(s): {sorted(missing_required)}. "
                f"Found keys: {all_keys}. "
                f"Matched: {detected_columns}."
            ],
        )

    items: List[ParsedBOMItem] = []
    skipped: List[SkippedRow] = []

    for row_idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            skipped.append(SkippedRow(
                row_number=row_idx,
                reason=f"Not a JSON object: {type(row).__name__}",
            ))
            continue

        item, skip = _parse_row(row, column_mapping, row_number=row_idx)
        if item:
            items.append(item)
        elif skip:
            skipped.append(skip)

    return BOMParseResult(
        items=items,
        skipped=skipped,
        column_mapping=column_mapping,
        detected_columns=detected_columns,
        unmatched_columns=unmatched_columns,
        total_rows=len(rows),
        warnings=parse_warnings,
    )


def parse_xlsx(content: bytes) -> BOMParseResult:
    """Parse an XLSX BOM file.

    Uses the first sheet, or a sheet named "BOM" or "Parts" if present.

    Args:
        content: Raw XLSX bytes.

    Returns:
        BOMParseResult with parsed items and metadata.
    """
    try:
        import openpyxl
    except ImportError:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=["openpyxl is required for XLSX parsing but not installed"],
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=[f"Failed to open XLSX: {exc}"],
        )

    # Select sheet: prefer "BOM" or "Parts", fall back to first sheet
    sheet = None
    for name in wb.sheetnames:
        if name.lower() in ("bom", "parts"):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active or wb[wb.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)

    # Read header row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=["Empty XLSX sheet"],
        )

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    # Filter out empty headers
    headers = [h for h in headers if h]

    if not headers:
        wb.close()
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=["No headers found in XLSX sheet"],
        )

    # Re-read with full header for proper column count
    all_headers = [str(h).strip() if h is not None else f"_col_{i}" for i, h in enumerate(header_row)]
    column_mapping, detected_columns, unmatched_columns = match_columns(all_headers)

    parse_warnings: List[str] = []
    missing_required = REQUIRED_COLUMNS - set(detected_columns)
    if missing_required:
        wb.close()
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping=column_mapping,
            detected_columns=detected_columns,
            unmatched_columns=unmatched_columns,
            total_rows=0,
            warnings=[
                f"Missing required column(s): {sorted(missing_required)}. "
                f"Found columns: {headers}. "
                f"Matched: {detected_columns}."
            ],
        )

    items: List[ParsedBOMItem] = []
    skipped: List[SkippedRow] = []
    total_rows = 0

    for row_idx, row_values in enumerate(rows_iter, start=2):
        total_rows += 1
        # Build row dict from header mapping
        row_dict: Dict[str, Any] = {}
        for col_idx, value in enumerate(row_values):
            if col_idx < len(all_headers):
                row_dict[all_headers[col_idx]] = value

        item, skip = _parse_row(row_dict, column_mapping, row_number=row_idx)
        if item:
            items.append(item)
        elif skip:
            skipped.append(skip)

    wb.close()

    return BOMParseResult(
        items=items,
        skipped=skipped,
        column_mapping=column_mapping,
        detected_columns=detected_columns,
        unmatched_columns=unmatched_columns,
        total_rows=total_rows,
        warnings=parse_warnings,
    )


def parse_bom_file(content: bytes, filename: str) -> BOMParseResult:
    """Parse a BOM file based on its filename extension.

    Args:
        content: Raw file bytes.
        filename: Original filename (used to detect format).

    Returns:
        BOMParseResult with parsed items and metadata.
    """
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return parse_csv(content)
    elif lower_name.endswith(".json"):
        return parse_json(content)
    elif lower_name.endswith((".xlsx", ".xls")):
        return parse_xlsx(content)
    else:
        return BOMParseResult(
            items=[],
            skipped=[],
            column_mapping={},
            detected_columns=[],
            unmatched_columns=[],
            total_rows=0,
            warnings=[
                f"Unsupported file format: {filename}. "
                "Supported formats: CSV, JSON, XLSX"
            ],
        )


# ---------------------------------------------------------------------------
# Material percentage computation
# ---------------------------------------------------------------------------
def compute_material_percentages(
    items: List[ParsedBOMItem],
) -> Dict[str, Dict[str, float]]:
    """Compute material percentages by weight and by value across BOM items.

    Args:
        items: List of parsed BOM items.

    Returns:
        Dict with 'by_weight' and 'by_value' keys, each mapping
        material name to percentage (0-100).
    """
    weight_totals: Dict[str, float] = {}
    value_totals: Dict[str, float] = {}
    total_weight = 0.0
    total_value = 0.0

    for item in items:
        mat = (item.material or "unknown").lower()

        if item.weight_kg is not None and item.weight_kg > 0:
            qty = item.quantity if item.quantity and item.quantity > 0 else 1.0
            w = item.weight_kg * qty
            weight_totals[mat] = weight_totals.get(mat, 0.0) + w
            total_weight += w

        if item.value_usd is not None and item.value_usd > 0:
            qty = item.quantity if item.quantity and item.quantity > 0 else 1.0
            v = item.value_usd * qty
            value_totals[mat] = value_totals.get(mat, 0.0) + v
            total_value += v

    by_weight: Dict[str, float] = {}
    if total_weight > 0:
        for mat, w in sorted(weight_totals.items()):
            by_weight[mat] = round(w / total_weight * 100, 2)

    by_value: Dict[str, float] = {}
    if total_value > 0:
        for mat, v in sorted(value_totals.items()):
            by_value[mat] = round(v / total_value * 100, 2)

    return {"by_weight": by_weight, "by_value": by_value}
