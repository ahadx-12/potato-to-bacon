"""Excel workbook generator for BOM engineering reports.

Produces a professional, client-deliverable Excel workbook from a
BOMEngineeringReport.  The workbook contains:

  Sheet 1: Executive Summary
    - Company info, analysis date, total SKUs
    - Annual duty exposure and achievable savings (if value/volume known)
    - Opportunity count by type
    - Risk summary

  Sheet 2: All Opportunities (ranked by annual savings)
    - One row per opportunity
    - All fields: SKU, type, confidence, baseline/optimized rates,
      annual savings, action items (condensed), legal basis

  Sheet 3: Quick Wins
    - Documentation-only and FTA utilization opportunities
    - Can be actioned immediately, no physical changes needed

  Sheet 4: Per-SKU Duty Breakdown
    - One row per SKU
    - Base rate, 232, 301, AD/CVD, FTA, exclusion, total effective
    - Opportunities count

  Sheet 5: Compliance Risk Findings
    - Risk category, severity, exposure, penalty modeling
    - Immediate actions

  Sheet 6: Implementation Roadmap
    - Opportunities grouped by implementation type
    - Sorted by payback and complexity

The workbook uses color coding and professional formatting suitable for
presentation to a company's CFO and trade compliance team.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------

_NAVY = "1A237E"        # Primary brand (dark blue)
_DARK_BLUE = "283593"
_ACCENT_BLUE = "0D47A1"
_LIGHT_BLUE = "E8EAF6"  # Header background (light)
_WHITE = "FFFFFF"
_LIGHT_GRAY = "F5F5F5"
_MID_GRAY = "9E9E9E"
_DARK_GRAY = "424242"

_GREEN = "2E7D32"       # Savings / good
_AMBER = "F57F17"       # Warning / medium
_RED = "C62828"         # Risk / critical
_TEAL = "00695C"        # Quick wins

_CRITICAL_BG = "FFEBEE"
_HIGH_BG = "FFF8E1"
_MEDIUM_BG = "F3F8FF"

# Severity colors
_SEVERITY_COLORS = {
    "critical": (_RED, _CRITICAL_BG),
    "high": ("E65100", _HIGH_BG),
    "medium": ("1565C0", _MEDIUM_BG),
    "low": (_MID_GRAY, _LIGHT_GRAY),
}

# Opportunity type colors
_OPP_TYPE_COLORS = {
    "documentation": (_TEAL, "E0F2F1"),
    "fta_utilization": (_GREEN, "E8F5E9"),
    "exclusion_filing": (_GREEN, "F1F8E9"),
    "reclassification": (_ACCENT_BLUE, "E3F2FD"),
    "product_engineering": ("6A1B9A", "F3E5F5"),
    "trade_lane": ("BF360C", "FBE9E7"),
    "ad_cvd_exposure": (_RED, _CRITICAL_BG),
    "ad_cvd_engineering": (_RED, "FCE4EC"),
}


def _make_font(*, bold=False, size=10, color=_DARK_GRAY, italic=False) -> "Font":
    return Font(bold=bold, size=size, color=color, italic=italic)


def _make_fill(hex_color: str) -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> "Border":
    thin = Side(style="thin", color=_MID_GRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_border() -> "Border":
    medium = Side(style="medium", color=_NAVY)
    return Border(left=medium, right=medium, top=medium, bottom=medium)


def _set_col_width(ws: Any, col_idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_header_row(ws: Any, row: int, num_cols: int) -> None:
    """Apply navy-on-white header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _make_font(bold=True, size=10, color=_WHITE)
        cell.fill = _make_fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _header_border()


def _write_row(ws: Any, row: int, values: List, bg_color: str = _WHITE,
               bold: bool = False, wrap: bool = True) -> None:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val if not isinstance(val, list) else "; ".join(str(v) for v in val[:3])
        cell.font = _make_font(bold=bold, size=9)
        cell.fill = _make_fill(bg_color)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        cell.border = _thin_border()


def _fmt_rate(rate: Optional[float]) -> str:
    if rate is None:
        return ""
    return f"{rate:.2f}%"


def _fmt_usd(amount: Optional[float]) -> str:
    if amount is None:
        return ""
    return f"${amount:,.0f}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_executive_summary(ws: Any, report: Any, company_name: str) -> None:
    """Sheet 1: Executive summary."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    # Title block
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "TARIFF ENGINEERING REPORT — EXECUTIVE SUMMARY"
    title.font = _make_font(bold=True, size=16, color=_WHITE)
    title.fill = _make_fill(_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"{company_name} | Analysis date: {report.analyzed_at[:10]}"
    sub.font = _make_font(size=11, color=_DARK_BLUE, italic=True)
    sub.fill = _make_fill(_LIGHT_BLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Key metrics
    row = 4
    ps = report.portfolio_summary

    def _kv(label: str, value: str, label_color: str = _NAVY) -> None:
        nonlocal row
        lc = ws.cell(row=row, column=1)
        lc.value = label
        lc.font = _make_font(bold=True, size=10, color=label_color)
        lc.fill = _make_fill(_LIGHT_BLUE)
        lc.border = _thin_border()

        vc = ws.cell(row=row, column=2)
        vc.value = value
        vc.font = _make_font(size=10)
        vc.border = _thin_border()
        row += 1

    ws.merge_cells(f"A{row}:H{row}")
    section = ws.cell(row=row, column=1)
    section.value = "PORTFOLIO OVERVIEW"
    section.font = _make_font(bold=True, size=11, color=_WHITE)
    section.fill = _make_fill(_DARK_BLUE)
    row += 1

    _kv("Total SKUs Analyzed", str(ps.skus_analyzed))
    _kv("SKUs With Opportunities", str(ps.skus_with_opportunities))
    _kv("SKUs With AD/CVD Exposure", str(ps.skus_with_adcvd_exposure))
    if ps.total_annual_duty_exposure is not None:
        _kv("Total Annual Duty Exposure", _fmt_usd(ps.total_annual_duty_exposure), _RED)
    if ps.achievable_annual_savings is not None:
        _kv("Achievable Annual Savings", _fmt_usd(ps.achievable_annual_savings), _GREEN)
    if ps.weighted_avg_baseline_rate is not None:
        _kv("Avg Effective Duty Rate (Baseline)", _fmt_rate(ps.weighted_avg_baseline_rate))
    if ps.weighted_avg_optimized_rate is not None:
        _kv("Avg Effective Duty Rate (Optimized)", _fmt_rate(ps.weighted_avg_optimized_rate))

    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    section2 = ws.cell(row=row, column=1)
    section2.value = "OPPORTUNITY COUNTS BY TYPE"
    section2.font = _make_font(bold=True, size=11, color=_WHITE)
    section2.fill = _make_fill(_DARK_BLUE)
    row += 1

    opp_types = [
        ("Documentation (Quick Wins)", ps.documentation_only_count),
        ("FTA Utilization", ps.fta_utilization_count),
        ("Exclusion Filing", ps.exclusion_filing_count),
        ("Reclassification", ps.reclassification_count),
        ("Product Engineering", ps.product_engineering_count),
        ("Trade Lane Optimization", ps.trade_lane_count),
        ("AD/CVD Exposure (Risk)", ps.adcvd_exposure_count),
    ]
    for label, count in opp_types:
        if count > 0:
            _kv(label, str(count))

    # Risk summary
    if report.risk_summary:
        rs = report.risk_summary
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        section3 = ws.cell(row=row, column=1)
        section3.value = "COMPLIANCE RISK SUMMARY"
        section3.font = _make_font(bold=True, size=11, color=_WHITE)
        section3.fill = _make_fill(_RED if rs.overall_risk_level in ("critical", "high") else _DARK_BLUE)
        row += 1

        _kv("Overall Risk Level", rs.overall_risk_level.upper())
        _kv("Total Risk Findings", str(rs.total_risk_findings))
        if rs.critical_count:
            _kv("Critical Findings", str(rs.critical_count), _RED)
        if rs.high_count:
            _kv("High Severity Findings", str(rs.high_count), "E65100")
        if rs.total_estimated_exposure_usd:
            _kv("Estimated Penalty Exposure", _fmt_usd(rs.total_potential_penalty_usd), _RED)
        if rs.prior_disclosure_recommended:
            _kv("Prior Disclosure to CBP", "RECOMMENDED", _RED)

    # Column widths
    _set_col_width(ws, 1, 38)
    _set_col_width(ws, 2, 22)
    for c in range(3, 9):
        _set_col_width(ws, c, 12)


def _build_opportunities_sheet(ws: Any, opportunities: List[Any], title: str) -> None:
    """Generic opportunities sheet builder."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title
    num_cols = 11
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    t = ws["A1"]
    t.value = title
    t.font = _make_font(bold=True, size=13, color=_WHITE)
    t.fill = _make_fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "SKU / Part ID", "Opportunity Type", "Confidence", "Risk Grade",
        "Current HTS", "Baseline Rate", "Optimized Rate", "Savings (pp)",
        "Annual Savings", "Summary", "Action Required"
    ]
    _write_row(ws, 2, headers)
    _style_header_row(ws, 2, num_cols)
    ws.row_dimensions[2].height = 22

    for r_idx, opp in enumerate(opportunities, start=3):
        opp_type = opp.opportunity_type if isinstance(opp.opportunity_type, str) else opp.opportunity_type.value
        conf = opp.confidence if isinstance(opp.confidence, str) else opp.confidence.value

        bg = _LIGHT_GRAY if r_idx % 2 == 0 else _WHITE
        type_colors = _OPP_TYPE_COLORS.get(opp_type, (_DARK_GRAY, bg))
        row_bg = type_colors[1] if opp.is_risk_finding else bg

        values = [
            opp.sku_id or "",
            opp_type.replace("_", " ").title(),
            conf.upper(),
            opp.risk_grade if isinstance(opp.risk_grade, str) else opp.risk_grade.value,
            opp.current_hts_code or "",
            _fmt_rate(opp.baseline_total_rate),
            _fmt_rate(opp.optimized_total_rate),
            _fmt_rate(opp.rate_reduction_pct),
            _fmt_usd(opp.annual_savings_estimate),
            opp.title[:100] if opp.title else "",
            opp.action_items[0][:80] if opp.action_items else "",
        ]
        _write_row(ws, r_idx, values, bg_color=row_bg)

        # Color the type cell
        type_cell = ws.cell(row=r_idx, column=2)
        fc, bg_c = type_colors
        type_cell.font = _make_font(bold=True, size=9, color=fc)

        # Color savings green if positive
        sav_cell = ws.cell(row=r_idx, column=9)
        if opp.annual_savings_estimate and opp.annual_savings_estimate > 0:
            sav_cell.font = _make_font(bold=True, size=9, color=_GREEN)
        elif opp.is_risk_finding:
            sav_cell.font = _make_font(bold=True, size=9, color=_RED)

    # Column widths
    widths = [16, 20, 12, 10, 12, 13, 13, 11, 14, 45, 45]
    for c_idx, w in enumerate(widths, start=1):
        _set_col_width(ws, c_idx, w)


def _build_sku_breakdown_sheet(ws: Any, sku_findings: List[Any]) -> None:
    """Sheet 4: Per-SKU duty breakdown."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    num_cols = 13
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    t = ws["A1"]
    t.value = "PER-SKU DUTY BREAKDOWN"
    t.font = _make_font(bold=True, size=13, color=_WHITE)
    t.fill = _make_fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "SKU / Part ID", "Description", "Origin", "HTS Code", "Category",
        "Base Rate", "Sec. 232", "Sec. 301", "AD Duty", "CVD Duty",
        "Excl. Relief", "Total Rate", "# Opportunities"
    ]
    _write_row(ws, 2, headers)
    _style_header_row(ws, 2, num_cols)
    ws.row_dimensions[2].height = 22

    for r_idx, sku in enumerate(sku_findings, start=3):
        bg = _LIGHT_GRAY if r_idx % 2 == 0 else _WHITE

        # Flag high-rate products
        total = sku.total_effective_rate
        if total >= 40.0:
            bg = _CRITICAL_BG
        elif total >= 25.0:
            bg = _HIGH_BG

        opp_count = len([o for o in sku.opportunities if not o.is_risk_finding])

        values = [
            sku.sku_id or "",
            (sku.description or "")[:60],
            sku.origin_country or "",
            sku.current_hts_code or "",
            sku.inferred_category or "",
            _fmt_rate(sku.base_rate),
            _fmt_rate(sku.section_232_rate) if sku.section_232_rate else "",
            _fmt_rate(sku.section_301_rate) if sku.section_301_rate else "",
            _fmt_rate(sku.ad_duty_rate) if sku.ad_duty_rate else "",
            _fmt_rate(sku.cvd_duty_rate) if sku.cvd_duty_rate else "",
            _fmt_rate(sku.exclusion_relief_rate) if sku.exclusion_relief_rate else "",
            _fmt_rate(sku.total_effective_rate),
            str(opp_count),
        ]
        _write_row(ws, r_idx, values, bg_color=bg)

        # Bold total rate
        total_cell = ws.cell(row=r_idx, column=12)
        total_cell.font = _make_font(bold=True, size=9,
                                     color=_RED if total >= 30.0 else _DARK_GRAY)

    widths = [14, 42, 8, 12, 14, 10, 9, 9, 9, 9, 11, 10, 15]
    for c_idx, w in enumerate(widths, start=1):
        _set_col_width(ws, c_idx, w)


def _build_risk_sheet(ws: Any, risk_findings: List[Any]) -> None:
    """Sheet 5: Compliance risk findings."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    num_cols = 10
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    t = ws["A1"]
    t.value = "COMPLIANCE RISK FINDINGS"
    t.font = _make_font(bold=True, size=13, color=_WHITE)
    t.fill = _make_fill(_RED)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if not risk_findings:
        ws.cell(row=3, column=1).value = "No compliance risk findings identified."
        ws.cell(row=3, column=1).font = _make_font(size=11, color=_GREEN)
        return

    headers = [
        "SKU / Part ID", "Category", "Severity", "Risk Summary",
        "Est. Annual Exposure", "Potential Penalty", "Confidence",
        "Immediate Action", "Prior Disclosure?", "Legal Counsel?"
    ]
    _write_row(ws, 2, headers)
    _style_header_row(ws, 2, num_cols)
    ws.row_dimensions[2].height = 22

    for r_idx, risk in enumerate(risk_findings, start=3):
        severity = risk.severity if isinstance(risk.severity, str) else risk.severity.value
        category = risk.category if isinstance(risk.category, str) else risk.category.value

        sev_fg, sev_bg = _SEVERITY_COLORS.get(severity, (_DARK_GRAY, _WHITE))

        values = [
            risk.sku_id or "Portfolio",
            category.replace("_", " ").title(),
            severity.upper(),
            risk.risk_summary[:100],
            _fmt_usd(risk.estimated_annual_exposure_usd),
            _fmt_usd(risk.potential_penalty_usd),
            risk.confidence.upper(),
            risk.immediate_actions[0][:80] if risk.immediate_actions else "",
            "YES" if risk.prior_disclosure_recommended else "No",
            "YES" if risk.requires_legal_counsel else "No",
        ]
        _write_row(ws, r_idx, values, bg_color=sev_bg)

        # Severity cell
        sev_cell = ws.cell(row=r_idx, column=3)
        sev_cell.font = _make_font(bold=True, size=9, color=sev_fg)

        # Highlight prior disclosure
        pd_cell = ws.cell(row=r_idx, column=9)
        if risk.prior_disclosure_recommended:
            pd_cell.font = _make_font(bold=True, size=9, color=_RED)

    widths = [14, 24, 12, 48, 18, 18, 12, 50, 16, 15]
    for c_idx, w in enumerate(widths, start=1):
        _set_col_width(ws, c_idx, w)


def _build_roadmap_sheet(ws: Any, opportunities: List[Any]) -> None:
    """Sheet 6: Implementation roadmap — grouped by type and sorted by payback."""

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    num_cols = 8
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    t = ws["A1"]
    t.value = "IMPLEMENTATION ROADMAP"
    t.font = _make_font(bold=True, size=13, color=_WHITE)
    t.fill = _make_fill(_DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Phase", "SKU / Part ID", "Opportunity", "Action Required",
        "Evidence Needed", "Annual Savings", "Payback (mo)", "Risk Grade"
    ]
    _write_row(ws, 2, headers)
    _style_header_row(ws, 2, num_cols)
    ws.row_dimensions[2].height = 22

    # Define implementation phases
    phase_order = [
        ("Phase 1 — Documentation Only (Act Now)", ["documentation"]),
        ("Phase 2 — FTA & Exclusion (Low Effort)", ["fta_utilization", "exclusion_filing"]),
        ("Phase 3 — Reclassification (Review Required)", ["reclassification"]),
        ("Phase 4 — Supply Chain & Engineering (Longer Term)", ["trade_lane", "product_engineering"]),
    ]

    savings_opps = [o for o in opportunities if not o.is_risk_finding]
    row_idx = 3

    for phase_label, phase_types in phase_order:
        phase_opps = [
            o for o in savings_opps
            if (o.opportunity_type if isinstance(o.opportunity_type, str)
                else o.opportunity_type.value) in phase_types
        ]
        if not phase_opps:
            continue

        # Phase header row
        ws.merge_cells(f"A{row_idx}:{get_column_letter(num_cols)}{row_idx}")
        ph = ws.cell(row=row_idx, column=1)
        ph.value = phase_label
        ph.font = _make_font(bold=True, size=10, color=_WHITE)
        ph.fill = _make_fill(_ACCENT_BLUE)
        ph.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

        # Sort by annual savings descending
        phase_opps_sorted = sorted(
            phase_opps,
            key=lambda o: -(o.annual_savings_estimate or 0.0)
        )

        for opp in phase_opps_sorted:
            opp_type = opp.opportunity_type if isinstance(opp.opportunity_type, str) else opp.opportunity_type.value
            payback = opp.payback_months
            bg = _LIGHT_GRAY if row_idx % 2 == 0 else _WHITE

            values = [
                phase_label.split("—")[0].strip(),
                opp.sku_id or "",
                opp.title[:60] if opp.title else "",
                opp.action_items[0][:70] if opp.action_items else "",
                opp.evidence_required[0][:60] if opp.evidence_required else "",
                _fmt_usd(opp.annual_savings_estimate),
                f"{payback:.1f}" if payback is not None else "—",
                opp.risk_grade if isinstance(opp.risk_grade, str) else opp.risk_grade.value,
            ]
            _write_row(ws, row_idx, values, bg_color=bg)

            sav_cell = ws.cell(row=row_idx, column=6)
            if opp.annual_savings_estimate and opp.annual_savings_estimate > 0:
                sav_cell.font = _make_font(bold=True, size=9, color=_GREEN)

            row_idx += 1

    widths = [22, 14, 45, 50, 50, 16, 14, 12]
    for c_idx, w in enumerate(widths, start=1):
        _set_col_width(ws, c_idx, w)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel_report(
    report: Any,
    company_name: str = "Importer of Record",
) -> bytes:
    """Generate an Excel workbook from a BOMEngineeringReport.

    Args:
        report: A BOMEngineeringReport dataclass instance.
        company_name: Name for the executive summary header.

    Returns:
        Raw bytes of the .xlsx file.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()

    # Remove default sheet
    default = wb.active
    if default:
        wb.remove(default)

    # Sheet 1: Executive Summary
    ws_exec = wb.create_sheet("Executive Summary")
    _build_executive_summary(ws_exec, report, company_name)

    # Sheet 2: All Opportunities
    all_opps = [o for o in report.all_opportunities if not o.is_risk_finding]
    if all_opps:
        ws_opps = wb.create_sheet("All Opportunities")
        _build_opportunities_sheet(ws_opps, all_opps, "ALL OPPORTUNITIES — RANKED BY ANNUAL SAVINGS")

    # Sheet 3: Quick Wins
    if report.quick_wins:
        ws_qw = wb.create_sheet("Quick Wins")
        _build_opportunities_sheet(ws_qw, report.quick_wins, "QUICK WINS — NO PHYSICAL CHANGES REQUIRED")

    # Sheet 4: Per-SKU Duty Breakdown
    if report.sku_findings:
        ws_sku = wb.create_sheet("Per-SKU Breakdown")
        _build_sku_breakdown_sheet(ws_sku, report.sku_findings)

    # Sheet 5: Compliance Risk Findings
    ws_risk = wb.create_sheet("Compliance Risks")
    _build_risk_sheet(ws_risk, report.risk_findings or [])

    # Sheet 6: Implementation Roadmap
    if all_opps:
        ws_road = wb.create_sheet("Implementation Roadmap")
        _build_roadmap_sheet(ws_road, report.all_opportunities)

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
