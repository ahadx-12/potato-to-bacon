"""Generate XLSX test fixtures for BOM parser tests."""

from pathlib import Path
import openpyxl


def create_multi_sheet_xlsx():
    """Create an XLSX with multiple sheets; parser should use 'BOM' sheet."""
    wb = openpyxl.Workbook()

    # Default sheet (should be ignored in favor of "BOM" sheet)
    ws_default = wb.active
    ws_default.title = "Summary"
    ws_default.append(["This is the summary sheet - should be ignored"])

    # BOM sheet - the one the parser should pick up
    ws_bom = wb.create_sheet("BOM")
    ws_bom.append(["Part ID", "Description", "Material", "Weight (kg)", "Unit Cost", "Country of Origin", "HTS Code"])
    ws_bom.append(["X-001", "Steel hex bolt M12x60", "Steel", 0.065, 1.15, "CN", "7318.15.20"])
    ws_bom.append(["X-002", "Aluminum spacer ring", "Aluminum", 0.030, 0.95, "TW", None])
    ws_bom.append(["X-003", "Rubber O-ring seal", "Rubber", 0.005, 0.20, "CN", None])
    ws_bom.append([None, None, None, None, None, None, None])  # Empty row
    ws_bom.append(["X-004", "Copper terminal lug", "Copper", 0.025, 1.80, "US", None])
    ws_bom.append(["X-005", "Nylon insulating sleeve", "Nylon", 0.010, 0.45, "CN", None])

    # Parts sheet (alternative naming)
    ws_parts = wb.create_sheet("Parts")
    ws_parts.append(["SKU", "Desc", "Mat", "Price"])
    ws_parts.append(["ALT-001", "Plastic housing cover", "ABS Plastic", 2.50])

    fixture_dir = Path(__file__).parent
    wb.save(fixture_dir / "multi_sheet_bom.xlsx")
    wb.close()


if __name__ == "__main__":
    create_multi_sheet_xlsx()
    print("Created multi_sheet_bom.xlsx")
